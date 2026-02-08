#!/usr/bin/env python3
"""Convert KWB ModbusInfo Excel files to JSON configuration format."""

import json
import logging
import re
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl library is required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class ModbusInfoConverter:
    """Converts ModbusInfo Excel files to JSON configuration."""

    # Character replacements for entity ID sanitization (same as coordinator.py)
    _ENTITY_ID_REPLACEMENTS = str.maketrans({
        " ": "_",
        ".": "_",
        "/": "_",
        "-": "_",
        ":": "_",
        "ä": "ae",
        "ö": "oe",
        "ü": "ue",
        "Ä": "ae",
        "Ö": "oe",
        "Ü": "ue",
        "ß": "ss",
        "&": "and",
        "@": "at",
        "(": "",
        ")": "",
        "#": "",
        "!": "",
        "?": "",
        ",": "",
        ";": "",
        "'": "",
        '"': "",
    })

    # Equipment prefixes that are 0-indexed (need +1 for entity_id)
    ZERO_INDEXED_PREFIXES = {"PUF", "BUF", "ZIR", "Circ", "WMZ", "HQM", "Zirk"}

    # Sheet categories
    #UNIVERSAL_SHEET = "Universal"
    # WMM Autonom is NOT a device - it contains universal Modbus lifetick registers
    UNIVERSAL_SHEETS = ["Universal"]

    DEVICE_SHEETS = [
        "KWB Easyfire", "KWB EF3", "KWB Multifire", "KWB Pelletfire+",
        "KWB Combifire", "KWB CF 2", "KWB CF 1", "KWB CF 1.5",
        "KWB EasyAir Plus"
    ]

    # Equipment sheet names (German and English)
    EQUIPMENT_SHEETS = {
        # German: English
        "Heizkreise": "Heating circuits",
        "Pufferspeicher": "Buffer storage tank",
        "Brauchwasserspeicher": "DHWC",
        "Zweitwärmequellen": "Secondary heating sources",
        "Zirkulation": "Circulation",
        "Solar": "Solar",
        "Kesselfolgeschaltung": "Boiler master-and-slave circuit",
        "Wärmemengenzähler": "Heat quantity meter",
        "Übergabestation": "Transfer station",
    }

    ALARMS_SHEET = "Alarms"
    VALUE_TABLES_SHEET = "ValueTables"

    # Use consistent English filenames for equipment
    EQUIPMENT_FILE_MAP = {
        "Heizkreise": "heating_circuits.json",
        "Heating circuits": "heating_circuits.json",
        "Pufferspeicher": "buffer_storage.json",
        "Buffer storage tank": "buffer_storage.json",
        "Brauchwasserspeicher": "dhw_storage.json",
        "DHWC": "dhw_storage.json",
        "Zweitwärmequellen": "secondary_heat_sources.json",
        "Secondary heating sources": "secondary_heat_sources.json",
        "Zirkulation": "circulation.json",
        "Circulation": "circulation.json",
        "Solar": "solar.json",
        "Kesselfolgeschaltung": "boiler_sequence.json",
        "Boiler master-and-slave circuit": "boiler_sequence.json",
        "Wärmemengenzähler": "heat_meters.json",
        "Heat quantity meter": "heat_meters.json",
        "Übergabestation": "transfer_station.json",
        "Transfer station": "transfer_station.json",
    }

    # Map device sheet names to filenames
    DEVICE_FILE_MAP = {
        "KWB Easyfire": "kwb_easyfire.json",
        "KWB EF3": "kwb_ef3.json",
        "KWB Multifire": "kwb_multifire.json",
        "KWB Pelletfire+": "kwb_pelletfire_plus.json",
        "KWB Combifire": "kwb_combifire.json",
        "KWB CF 2": "kwb_cf2.json",
        "KWB CF 1": "kwb_cf1.json",
        "KWB CF 1.5": "kwb_cf1_5.json",
        "KWB EasyAir Plus": "kwb_easyair_plus.json",
    }

    def __init__(self, input_dir: Path, output_dir: Path):
        """Initialize converter."""
        self.input_dir = Path(input_dir)
        self.output_dir = Path(output_dir)
        # Cache for English data lookup: {version: {starting_address: {"name": str, "index": str}}}
        self._english_data_cache: dict[str, dict[int, dict[str, str]]] = {}

    def sanitize_for_entity_id(self, name: str) -> str:
        """Sanitize a string for use in Home Assistant entity IDs.

        This mirrors the sanitization logic in coordinator.py to ensure
        entity IDs are valid and consistent.
        """
        if not name:
            return ""

        # Apply character replacements
        result = name.translate(self._ENTITY_ID_REPLACEMENTS)

        # Convert to lowercase
        result = result.lower()

        # Remove any remaining invalid characters
        result = re.sub(r'[^a-z0-9_]', '', result)

        # Collapse multiple underscores
        result = re.sub(r'_+', '_', result)

        # Strip leading/trailing underscores
        result = result.strip('_')

        return result

    def parse_filename(self, filename: str) -> dict[str, str]:
        """Parse ModbusInfo filename to extract version and language."""
        match = re.match(r'ModbusInfo-(\w+)-V(\d+\.\d+\.\d+)\.xlsx', filename, re.IGNORECASE)
        if match:
            language = match.group(1).lower()
            version = match.group(2)
            return {"language": language, "version": version}
        return {}

    def load_english_data(self, version: str) -> dict[int, dict[str, str]]:
        """Load English register data (name and index) for a specific version.

        Returns a mapping of starting_address -> {"name": str, "index": str} for use
        in generating language-independent entity IDs.
        """
        if version in self._english_data_cache:
            return self._english_data_cache[version]

        english_data: dict[int, dict[str, str]] = {}
        en_file = self.input_dir / f"ModbusInfo-en-V{version}.xlsx"

        if not en_file.exists():
            logger.warning(f"English file not found for version {version}: {en_file}")
            self._english_data_cache[version] = english_data
            return english_data

        try:
            workbook = openpyxl.load_workbook(en_file, data_only=True)

            # Collect data from all relevant sheets
            all_sheets = (
                self.UNIVERSAL_SHEETS +
                self.DEVICE_SHEETS +
                list(self.EQUIPMENT_SHEETS.keys()) +
                list(self.EQUIPMENT_SHEETS.values())
            )

            for sheet_name in all_sheets:
                if sheet_name not in workbook.sheetnames:
                    continue

                sheet = workbook[sheet_name]
                headers = [cell.value for cell in sheet[1]]

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if not any(row) or row[0] is None:
                        continue

                    # Get StartingAddress, Name, and Index
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(headers) and headers[i]:
                            row_dict[headers[i]] = value

                    address = row_dict.get("StartingAddress")
                    name = row_dict.get("Name")
                    index = row_dict.get("Index")

                    if address is not None and name:
                        data = {"name": str(name).strip()}
                        if index:
                            data["index"] = str(index).strip()
                        english_data[int(address)] = data

            workbook.close()
            logger.info(f"  Loaded {len(english_data)} English entries for version {version}")

        except Exception as exc:
            logger.error(f"Error loading English data for version {version}: {exc}")

        self._english_data_cache[version] = english_data
        return english_data

    def read_register_sheet(
        self,
        workbook: openpyxl.Workbook,
        sheet_name: str,
        english_data: dict[int, dict[str, str]] | None = None
    ) -> list[dict]:
        """Read register data from Excel sheet.

        Args:
            workbook: The Excel workbook
            sheet_name: Name of the sheet to read
            english_data: Optional mapping of starting_address -> {"name": str, "index": str}
                          for generating language-independent entity IDs
        """
        if sheet_name not in workbook.sheetnames:
            return []

        sheet = workbook[sheet_name]
        registers = []

        # Get header row
        headers = [cell.value for cell in sheet[1]]

        # Read data rows (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row) or row[0] is None:  # Skip empty rows or rows without address
                continue

            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = value

            register = self.normalize_register(row_dict, english_data)
            if register:
                registers.append(register)

        return registers

    def _sanitize_index_for_entity_id(self, index: str) -> str | None:
        """Sanitize an index string for use in entity_id.

        Converts index like "SOL 1", "PUF 0", "HK 1.1" to entity_id format
        like "sol_1", "puf_1", "hk_1_1". Handles 0-indexed equipment by converting to 1-based.

        Args:
            index: The index string (e.g., "SOL 1", "PUF 0", "HK 1.1")

        Returns:
            Sanitized index string like "sol_1", "hk_1_1", or None if not applicable
        """
        if not index:
            return None

        # Split into prefix and number part
        parts = index.split(" ", 1)
        if len(parts) != 2:
            return None

        prefix = parts[0].strip()
        number_part = parts[1].strip()

        # Convert 0-indexed equipment to 1-based
        if prefix in self.ZERO_INDEXED_PREFIXES:
            try:
                num = int(number_part)
                number_part = str(num + 1)
            except ValueError:
                pass  # Keep original if not a simple number

        # Replace dots with underscores for heating circuits (HK 1.1 -> hk_1_1)
        sanitized_number = number_part.replace(".", "_")

        # Sanitize the prefix (lowercase, handle special chars)
        sanitized_prefix = self.sanitize_for_entity_id(prefix)

        return f"{sanitized_prefix}_{sanitized_number}"

    def normalize_register(
        self,
        data: dict,
        english_data: dict[int, dict[str, str]] | None = None
    ) -> dict | None:
        """Normalize register data to standard format.

        Args:
            data: Raw register data from Excel row
            english_data: Optional mapping of starting_address -> {"name": str, "index": str}
                          for generating language-independent entity IDs
        """
        # Skip if no address
        address = data.get("StartingAddress")
        if address is None:
            return None

        address_int = int(address)
        name = str(data.get("Name", "")).strip() if data.get("Name") else ""

        # Get English data for language-independent entity_id
        en_data = english_data.get(address_int, {}) if english_data else {}
        english_name = en_data.get("name", name)
        english_index = en_data.get("index", "")

        # Generate entity_id from English name (language-independent)
        base_entity_id = self.sanitize_for_entity_id(english_name)

        # Include English index in entity_id if present
        if english_index:
            index_prefix = self._sanitize_index_for_entity_id(english_index)
            if index_prefix:
                entity_id = f"{index_prefix}_{base_entity_id}"
            else:
                entity_id = base_entity_id
        else:
            entity_id = base_entity_id

        # Build normalized register
        register = {
            "starting_address": address_int,
            "name": name,
            "entity_id": entity_id,
            "data_type": self._parse_function_code(data.get("Functions")),
            "type": str(data.get("Type", "u16")).strip().lower(),
            "user_level": self._parse_access_level(data.get("UserLevel")),
            "expert_level": self._parse_access_level(data.get("ExpertLevel")),
        }

        # Keep the original index field for display purposes
        if data.get("Index"):
            register["index"] = str(data.get("Index")).strip()

        if data.get("Unit/ValueTable"):
            unit_or_table = str(data.get("Unit/ValueTable")).strip()
            register["unit_value_table"] = unit_or_table

        if data.get("Min") is not None:
            register["min"] = data.get("Min")

        if data.get("Max") is not None:
            register["max"] = data.get("Max")

        if data.get("NumberOfRegisters"):
            num_regs = data.get("NumberOfRegisters")
            if num_regs and num_regs != 1:
                register["number_of_registers"] = int(num_regs)

        if data.get("ID"):
            register["id"] = str(data.get("ID"))

        if data.get("Parameter"):
            register["parameter"] = str(data.get("Parameter"))

        # Remove empty string values
        register = {k: v for k, v in register.items() if v not in [None, ""]}

        return register if register.get("starting_address") and register.get("name") else None

    def _parse_function_code(self, functions: Any) -> str:
        """Parse Modbus function code(s)."""
        if not functions:
            return "04"

        func_str = str(functions).strip()

        # Check for common function codes
        if "04" in func_str:
            return "04"  # Input registers (read-only)
        if "03" in func_str:
            return "03"  # Holding registers (read-write potential)

        return "04"  # Default to input registers

    def _parse_access_level(self, access: Any) -> str:
        """Parse access level."""
        if not access:
            return "read"

        access_str = str(access).lower().strip()

        if "readwrite" in access_str or "rw" in access_str:
            return "readwrite"
        elif "write" in access_str:
            return "write"
        elif "read" in access_str:
            return "read"

        return "read"

    def read_value_tables(self, workbook: openpyxl.Workbook) -> dict:
        """Read value tables from Excel."""
        if self.VALUE_TABLES_SHEET not in workbook.sheetnames:
            return {}

        sheet = workbook[self.VALUE_TABLES_SHEET]
        value_tables = {}

        # Read rows (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row) or row[0] is None:
                continue

            table_name = str(row[0]).strip() if row[0] else None
            value = row[1]
            translation = str(row[2]).strip() if len(row) > 2 and row[2] else None

            if table_name and value is not None and translation:
                if table_name not in value_tables:
                    value_tables[table_name] = {}
                value_tables[table_name][str(value)] = translation

        return value_tables

    def read_alarm_codes(self, workbook: openpyxl.Workbook) -> list[dict]:
        """Read alarm codes from Excel."""
        if self.ALARMS_SHEET not in workbook.sheetnames:
            return []

        sheet = workbook[self.ALARMS_SHEET]
        alarm_codes = []

        # Read rows (skip header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row) or row[0] is None:
                continue

            alarm = {
                "starting_address": int(row[0]) if row[0] is not None else None,
                "function_code": str(row[1]).strip() if len(row) > 1 and row[1] else "02",
                "alarm_id": str(row[2]).strip() if len(row) > 2 and row[2] else "",
                "description": str(row[3]).strip() if len(row) > 3 and row[3] else "",
            }

            if alarm["starting_address"] is not None:
                alarm_codes.append(alarm)

        return alarm_codes

    def merge_registers(self, base_registers: list[dict], override_registers: list[dict]) -> list[dict]:
        """Merge two register lists, where override_registers take precedence by address.

        Args:
            base_registers: Base register list (e.g., from Combifire)
            override_registers: Override register list (e.g., from CF 1)

        Returns:
            Merged register list with overrides applied
        """
        # Create a dictionary indexed by starting_address from base
        merged = {reg["starting_address"]: reg for reg in base_registers}

        # Override with specific registers
        for reg in override_registers:
            merged[reg["starting_address"]] = reg

        # Return sorted by address
        return sorted(merged.values(), key=lambda x: x["starting_address"])

    def convert_file(self, xlsx_file: Path) -> None:
        """Convert a single ModbusInfo Excel file to JSON."""
        logger.info(f"Converting {xlsx_file.name}...")

        # Parse filename
        file_info = self.parse_filename(xlsx_file.name)
        if not file_info:
            logger.error(f"Could not parse filename: {xlsx_file.name}")
            return

        version = file_info["version"]
        language = file_info["language"]

        # Load English names for entity_id generation (language-independent IDs)
        english_data = self.load_english_data(version)

        # Create output directory structure
        version_dir = self.output_dir / f"v{version}" / language
        version_dir.mkdir(parents=True, exist_ok=True)

        # Create subdirectories
        devices_dir = version_dir / "devices"
        equipment_dir = version_dir / "equipment"
        devices_dir.mkdir(exist_ok=True)
        equipment_dir.mkdir(exist_ok=True)

        # Load workbook
        try:
            workbook = openpyxl.load_workbook(xlsx_file, data_only=True)
        except Exception as exc:
            logger.error(f"Error loading workbook {xlsx_file}: {exc}")
            return

        # Read universal registers from all universal sheets
        universal_registers = []
        for sheet_name in self.UNIVERSAL_SHEETS:
            if sheet_name in workbook.sheetnames:
                logger.info(f"  Reading {sheet_name} sheet...")
                registers = self.read_register_sheet(workbook, sheet_name, english_data)
                universal_registers.extend(registers)
                logger.info(f"    Found {len(registers)} registers")

        # Save universal registers
        modbus_registers_file = version_dir / "modbus_registers.json"
        with open(modbus_registers_file, 'w', encoding='utf-8') as f:
            json.dump({
                "universal_registers": universal_registers
            }, f, indent=2, ensure_ascii=False)
        logger.info(f"  Created modbus_registers.json with {len(universal_registers)} universal registers")

        # Read Combifire base registers (if exists) - applies to CF 1, CF 1.5, CF 2
        combifire_base_registers = []
        if "KWB Combifire" in workbook.sheetnames:
            logger.info(f"  Reading KWB Combifire sheet (base for CF models)...")
            combifire_base_registers = self.read_register_sheet(workbook, "KWB Combifire", english_data)
            logger.info(f"    Found {len(combifire_base_registers)} base registers")

        # CF models that inherit from Combifire
        cf_models = ["KWB CF 1", "KWB CF 1.5", "KWB CF 2"]

        # Read and save device-specific registers
        for sheet_name in self.DEVICE_SHEETS:
            # Skip Combifire in main loop - it's handled separately below
            if sheet_name == "KWB Combifire":
                continue

            if sheet_name in workbook.sheetnames:
                logger.info(f"  Reading {sheet_name} sheet...")
                device_registers = self.read_register_sheet(workbook, sheet_name, english_data)

                # Special handling for CF models: merge with Combifire base
                if sheet_name in cf_models and combifire_base_registers:
                    logger.info(f"    Merging {len(device_registers)} specific registers with {len(combifire_base_registers)} Combifire base registers...")
                    device_registers = self.merge_registers(combifire_base_registers, device_registers)
                    logger.info(f"    Result: {len(device_registers)} total registers")

                if device_registers:
                    filename = self.DEVICE_FILE_MAP.get(sheet_name, f"{sheet_name.lower().replace(' ', '_')}.json")
                    device_file = devices_dir / filename
                    with open(device_file, 'w', encoding='utf-8') as f:
                        json.dump({
                            "registers": device_registers
                        }, f, indent=2, ensure_ascii=False)
                    logger.info(f"    Created devices/{filename} with {len(device_registers)} registers")

        # Save Combifire as its own device file too (if it exists)
        if combifire_base_registers:
            combifire_file = devices_dir / "kwb_combifire.json"
            with open(combifire_file, 'w', encoding='utf-8') as f:
                json.dump({
                    "registers": combifire_base_registers
                }, f, indent=2, ensure_ascii=False)
            logger.info(f"    Created devices/kwb_combifire.json with {len(combifire_base_registers)} registers")

        # Read and save equipment-specific registers
        # Check for both German and English sheet names
        all_equipment_sheets = list(self.EQUIPMENT_SHEETS.keys()) + list(self.EQUIPMENT_SHEETS.values())
        processed_equipment = set()  # Track which equipment we've already processed

        for sheet_name in all_equipment_sheets:
            if sheet_name in workbook.sheetnames:
                # Get the English filename (consistent for both German and English)
                filename = self.EQUIPMENT_FILE_MAP.get(sheet_name)
                if filename and filename not in processed_equipment:
                    logger.info(f"  Reading {sheet_name} sheet...")
                    equipment_registers = self.read_register_sheet(workbook, sheet_name, english_data)
                    if equipment_registers:
                        equipment_file = equipment_dir / filename
                        with open(equipment_file, 'w', encoding='utf-8') as f:
                            json.dump({
                                "registers": equipment_registers
                            }, f, indent=2, ensure_ascii=False)
                        logger.info(f"    Created equipment/{filename} with {len(equipment_registers)} registers")
                        processed_equipment.add(filename)

        # Read and save value tables
        logger.info(f"  Reading {self.VALUE_TABLES_SHEET}...")
        value_tables = self.read_value_tables(workbook)
        value_tables_file = version_dir / "value_tables.json"
        with open(value_tables_file, 'w', encoding='utf-8') as f:
            json.dump({
                "value_tables": value_tables
            }, f, indent=2, ensure_ascii=False)
        logger.info(f"  Created value_tables.json with {len(value_tables)} tables")

        # Read and save alarm codes
        logger.info(f"  Reading {self.ALARMS_SHEET}...")
        alarm_codes = self.read_alarm_codes(workbook)
        alarm_codes_file = version_dir / "alarm_codes.json"
        with open(alarm_codes_file, 'w', encoding='utf-8') as f:
            json.dump({
                "alarm_codes": alarm_codes
            }, f, indent=2, ensure_ascii=False)
        logger.info(f"  Created alarm_codes.json with {len(alarm_codes)} alarms")

        logger.info(f"✓ Successfully converted {xlsx_file.name}\n")

    def convert_all(self) -> None:
        """Convert all ModbusInfo Excel files in input directory."""
        xlsx_files = list(self.input_dir.glob("ModbusInfo*.xlsx"))

        if not xlsx_files:
            logger.error(f"No ModbusInfo*.xlsx files found in {self.input_dir}")
            return

        logger.info(f"Found {len(xlsx_files)} files to convert\n")

        for xlsx_file in xlsx_files:
            try:
                self.convert_file(xlsx_file)
            except Exception as exc:
                logger.error(f"Error converting {xlsx_file.name}: {exc}")
                import traceback
                traceback.print_exc()
                continue

        logger.info("="*60)
        logger.info("Conversion complete!")
        logger.info(f"Output directory: {self.output_dir}")


def main():
    """Main entry point."""
    script_dir = Path(__file__).parent
    input_dir = script_dir / "modbusinfo"
    output_dir = script_dir / "config" / "versions"

    if not input_dir.exists():
        logger.error(f"Input directory not found: {input_dir}")
        sys.exit(1)

    logger.info("KWB ModbusInfo to JSON Converter")
    logger.info(f"Input directory: {input_dir}")
    logger.info(f"Output directory: {output_dir}")
    logger.info("="*60)
    logger.info("")

    converter = ModbusInfoConverter(input_dir, output_dir)
    converter.convert_all()


if __name__ == "__main__":
    main()
