#!/usr/bin/env python3
"""Generate ModbusInfo Excel files from JSON configuration files.

This script creates Excel files from existing JSON configs, useful when
the original Excel source files are not available for a version.
"""

import json
import logging
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl library is required. Install with: pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


# Map JSON filenames to Excel sheet names
DEVICE_SHEET_MAP = {
    "kwb_easyfire.json": "KWB Easyfire",
    "kwb_ef3.json": "KWB EF3",
    "kwb_multifire.json": "KWB Multifire",
    "kwb_pelletfire_plus.json": "KWB Pelletfire+",
    "kwb_combifire.json": "KWB Combifire",
    "kwb_cf2.json": "KWB CF 2",
    "kwb_cf1.json": "KWB CF 1",
    "kwb_cf1_5.json": "KWB CF 1.5",
    "kwb_easyair_plus.json": "KWB EasyAir Plus",
}

EQUIPMENT_SHEET_MAP_EN = {
    "heating_circuits.json": "Heating circuits",
    "buffer_storage.json": "Buffer storage tank",
    "dhw_storage.json": "DHWC",
    "secondary_heat_sources.json": "Secondary heating sources",
    "circulation.json": "Circulation",
    "solar.json": "Solar",
    "boiler_sequence.json": "Boiler master-and-slave circuit",
    "heat_meters.json": "Heat quantity meter",
    "transfer_station.json": "Transfer station",
}

EQUIPMENT_SHEET_MAP_DE = {
    "heating_circuits.json": "Heizkreise",
    "buffer_storage.json": "Pufferspeicher",
    "dhw_storage.json": "Brauchwasserspeicher",
    "secondary_heat_sources.json": "Zweitwärmequellen",
    "circulation.json": "Zirkulation",
    "solar.json": "Solar",
    "boiler_sequence.json": "Kesselfolgeschaltung",
    "heat_meters.json": "Wärmemengenzähler",
    "transfer_station.json": "Übergabestation",
}

# Excel column headers
REGISTER_HEADERS = [
    "StartingAddress",
    "Name",
    "Functions",
    "Type",
    "UserLevel",
    "ExpertLevel",
    "Index",
    "Unit/ValueTable",
    "Min",
    "Max",
    "NumberOfRegisters",
    "ID",
    "Parameter",
]

VALUE_TABLE_HEADERS = ["TableName", "Value", "Translation"]
ALARM_HEADERS = ["StartingAddress", "FunctionCode", "AlarmID", "Description"]


def create_styled_workbook():
    """Create a new workbook with styling."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    return wb


def add_header_row(sheet, headers):
    """Add styled header row to sheet."""
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill


def convert_register_to_row(register: dict) -> list:
    """Convert a register dict to Excel row values."""
    # Map data_type back to Functions format
    data_type = register.get("data_type", "04")
    if data_type == "04":
        functions = "04"
    elif data_type == "03":
        functions = "03/06"
    else:
        functions = data_type

    # Map access levels back
    user_level = register.get("user_level", "read")
    expert_level = register.get("expert_level", "read")

    if user_level == "readwrite":
        user_level = "ReadWrite"
    elif user_level == "write":
        user_level = "Write"
    else:
        user_level = "Read"

    if expert_level == "readwrite":
        expert_level = "ReadWrite"
    elif expert_level == "write":
        expert_level = "Write"
    else:
        expert_level = "Read"

    return [
        register.get("starting_address"),
        register.get("name"),
        functions,
        register.get("type", "s16").upper(),
        user_level,
        expert_level,
        register.get("index", ""),
        register.get("unit_value_table", ""),
        register.get("min"),
        register.get("max"),
        register.get("number_of_registers", 1),
        register.get("id", ""),
        register.get("parameter", ""),
    ]


def add_registers_to_sheet(sheet, registers: list[dict]):
    """Add registers to a sheet."""
    add_header_row(sheet, REGISTER_HEADERS)

    for row_idx, register in enumerate(registers, 2):
        row_values = convert_register_to_row(register)
        for col_idx, value in enumerate(row_values, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


def load_json_file(path: Path) -> dict | None:
    """Load a JSON file."""
    try:
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.warning(f"Could not load {path}: {e}")
        return None


def create_excel_from_json(version_dir: Path, output_file: Path, language: str):
    """Create an Excel file from JSON configs for a specific language."""
    lang_dir = version_dir / language
    if not lang_dir.exists():
        logger.error(f"Language directory not found: {lang_dir}")
        return False

    wb = create_styled_workbook()
    sheets_created = 0

    # Add Universal sheet from modbus_registers.json
    modbus_file = lang_dir / "modbus_registers.json"
    if modbus_file.exists():
        data = load_json_file(modbus_file)
        if data and "universal_registers" in data:
            sheet = wb.create_sheet("Universal")
            add_registers_to_sheet(sheet, data["universal_registers"])
            sheets_created += 1
            logger.info(f"  Created Universal sheet with {len(data['universal_registers'])} registers")

    # Add device sheets
    devices_dir = lang_dir / "devices"
    if devices_dir.exists():
        for json_file, sheet_name in DEVICE_SHEET_MAP.items():
            file_path = devices_dir / json_file
            if file_path.exists():
                data = load_json_file(file_path)
                if data and "registers" in data:
                    sheet = wb.create_sheet(sheet_name)
                    add_registers_to_sheet(sheet, data["registers"])
                    sheets_created += 1
                    logger.info(f"  Created {sheet_name} sheet with {len(data['registers'])} registers")

    # Add equipment sheets
    equipment_dir = lang_dir / "equipment"
    sheet_map = EQUIPMENT_SHEET_MAP_EN if language == "en" else EQUIPMENT_SHEET_MAP_DE

    if equipment_dir.exists():
        for json_file, sheet_name in sheet_map.items():
            file_path = equipment_dir / json_file
            if file_path.exists():
                data = load_json_file(file_path)
                if data and "registers" in data:
                    sheet = wb.create_sheet(sheet_name)
                    add_registers_to_sheet(sheet, data["registers"])
                    sheets_created += 1
                    logger.info(f"  Created {sheet_name} sheet with {len(data['registers'])} registers")

    # Add ValueTables sheet
    value_tables_file = lang_dir / "value_tables.json"
    if value_tables_file.exists():
        data = load_json_file(value_tables_file)
        if data and "value_tables" in data:
            sheet = wb.create_sheet("ValueTables")
            add_header_row(sheet, VALUE_TABLE_HEADERS)
            row_idx = 2
            for table_name, values in data["value_tables"].items():
                for value, translation in values.items():
                    sheet.cell(row=row_idx, column=1, value=table_name)
                    sheet.cell(row=row_idx, column=2, value=value)
                    sheet.cell(row=row_idx, column=3, value=translation)
                    row_idx += 1
            sheets_created += 1
            logger.info(f"  Created ValueTables sheet")

    # Add Alarms sheet
    alarm_file = lang_dir / "alarm_codes.json"
    if alarm_file.exists():
        data = load_json_file(alarm_file)
        if data and "alarm_codes" in data:
            sheet = wb.create_sheet("Alarms")
            add_header_row(sheet, ALARM_HEADERS)
            for row_idx, alarm in enumerate(data["alarm_codes"], 2):
                sheet.cell(row=row_idx, column=1, value=alarm.get("starting_address"))
                sheet.cell(row=row_idx, column=2, value=alarm.get("function_code"))
                sheet.cell(row=row_idx, column=3, value=alarm.get("alarm_id"))
                sheet.cell(row=row_idx, column=4, value=alarm.get("description"))
            sheets_created += 1
            logger.info(f"  Created Alarms sheet with {len(data['alarm_codes'])} alarms")

    if sheets_created > 0:
        wb.save(output_file)
        logger.info(f"  Saved {output_file.name} with {sheets_created} sheets")
        return True
    else:
        logger.warning(f"  No sheets created for {output_file.name}")
        return False


def process_version(version_dir: Path, output_dir: Path):
    """Process a version directory and create Excel files."""
    version = version_dir.name  # e.g., "v24.7.1"
    version_num = version[1:]   # e.g., "24.7.1"

    logger.info(f"\nProcessing {version}...")

    # Create English Excel file
    en_output = output_dir / f"ModbusInfo-en-V{version_num}.xlsx"
    if (version_dir / "en").exists():
        logger.info(f"  Creating English Excel file...")
        create_excel_from_json(version_dir, en_output, "en")

    # Create German Excel file
    de_output = output_dir / f"ModbusInfo-de-V{version_num}.xlsx"
    if (version_dir / "de").exists():
        logger.info(f"  Creating German Excel file...")
        create_excel_from_json(version_dir, de_output, "de")


def main():
    """Main entry point."""
    import sys

    script_dir = Path(__file__).parent
    config_dir = script_dir.parent / "custom_components" / "kwb_heating" / "config" / "versions"
    output_dir = script_dir / "modbusinfo"

    # Allow specifying a specific version
    if len(sys.argv) > 1:
        version = sys.argv[1]
        if not version.startswith("v"):
            version = f"v{version}"
        version_dirs = [config_dir / version]
    else:
        # Process all versions
        version_dirs = sorted(config_dir.glob("v*"))

    logger.info("JSON to Excel Converter")
    logger.info(f"Config directory: {config_dir}")
    logger.info(f"Output directory: {output_dir}")
    logger.info("=" * 60)

    output_dir.mkdir(parents=True, exist_ok=True)

    for version_dir in version_dirs:
        if not version_dir.is_dir():
            continue
        process_version(version_dir, output_dir)

    logger.info("\n" + "=" * 60)
    logger.info("Done!")
    logger.info(f"\nGenerated Excel files are in: {output_dir}")
    logger.info("You can now run convert_modbusinfo.py to regenerate JSON with consistent entity_ids")


if __name__ == "__main__":
    main()
